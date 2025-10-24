"""
class Employee:
    pass

emp = Employee()

"""
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title  = "MySheet"

ws['B9'] = "Hello World"
ws['D9'] = 85250
ws["F9"] = datetime.now()
ws["H9"].value = '=column()'

wb.save("FirstExcel.xlsx")